from openpyxl import Workbook

def build_workbook():
    wb = Workbook(); wb.remove(wb.active)
    cfg = wb.create_sheet("config")
    cfg["B1"] = "科目"
    subjects = [(1,"语文","语","否"),(2,"数学","数","否"),(3,"英语","英","否"),(4,"体育","体","是"),(5,"物理","物","否")]
    for i,(code,name,short,outdoor) in enumerate(subjects):
        r=2+i
        cfg.cell(row=r,column=1,value=code); cfg.cell(row=r,column=2,value=name)
        cfg.cell(row=r,column=3,value=short); cfg.cell(row=r,column=4,value=outdoor)
    # 时间线：同一天的时间流（上午+下午）放在同一条时间线里，代号=星期几
    # 时间线1「周一」(代号1)：行2=名称，行3=代号，行4起=条目
    cfg.cell(row=2,column=5,value="周一"); cfg.cell(row=3,column=5,value=1)
    monday = [("08:00",45,1),("08:55",45,2),("09:50",45,3),("14:00",45,4),("14:55",45,5)]
    for i,(s,d,df) in enumerate(monday):
        r=4+i; cfg.cell(row=r,column=5,value=s); cfg.cell(row=r,column=6,value=d); cfg.cell(row=r,column=7,value=df)
    # 时间线2「周二」(代号2)
    cfg.cell(row=2,column=8,value="周二"); cfg.cell(row=3,column=8,value=2)
    tuesday = [("08:00",45,1),("08:55",45,2),("09:50",45,3),("14:00",45,4),("14:55",45,5)]
    for i,(s,d,df) in enumerate(tuesday):
        r=4+i; cfg.cell(row=r,column=8,value=s); cfg.cell(row=r,column=9,value=d); cfg.cell(row=r,column=10,value=df)
    # 杂项配置
    cfg.cell(row=2,column=28,value="classisland.theme"); cfg.cell(row=2,column=29,value='"Light"')
    cfg.cell(row=3,column=28,value="classwidgets.tts_enabled"); cfg.cell(row=3,column=29,value="true")
    # 年级 sheet：行1 班级名；A列 课位"周几-第几节"
    g1 = wb.create_sheet("高一")
    g1["B1"]="1班"; g1["C1"]="2班"
    slots=[("周一-1",2,1),("周一-2",1,2),("周一-3",3,1),("周一-4",4,5),("周一-5",2,4),
           ("周二-1",1,2),("周二-2",3,3),("周二-3",5,4),("周二-4",4,1),("周二-5",2,5)]
    for i,(slot,c1,c2) in enumerate(slots):
        r=2+i; g1.cell(row=r,column=1,value=slot); g1.cell(row=r,column=2,value=c1); g1.cell(row=r,column=3,value=c2)
    return wb


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "samples/sample_timetable.xlsx"
    build_workbook().save(out)
    print("示例已生成:", out)
