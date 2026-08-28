"""Excel 解析器。

格式约定（见 plan.md）：
- 每个非 config 的 sheet = 一个年级。
- 年级 sheet 布局：
    * 行 1 从 B1 起为班级名称（B1、C1、D1 …）。
    * A 列从行 2 起为课位标记 "周几-第几节"（如 "周一-1"），
      对应行 B 列起为 config sheet 中的科目代号。
- config sheet：
    * A 列科目代号（行 2 起 1..n），B 列科目名，C 列简称，D 列是否室外课<是/否>。
    * E..Z 为时间线区域：行 2 是时间线名称，行 3 为代号（可选），行 4 起为条目；
      每条时间线占三列：<开始时间 hh:mm, 持续分钟, 默认课程>；
      同一天的时间流（上午/下午/晚自习）全部放在同一条时间线里，
      时间线代号对应星期几（1=周一 … 7=周日）；相邻条目间的空隙自动视为课间；
      最多七条时间线，留空不启用。
    * AB 列（行 2 起）为杂项配置名称，AC 列为配置内容。
"""
import re
from typing import Dict, List, Optional

from openpyxl import load_workbook

DAY_ORDER = {"周一": 1, "周二": 2, "周三": 3, "周四": 4, "周五": 5, "周六": 6, "周日": 7}

_SLOT_RE = re.compile(r"^\s*(周[一二三四五六日天])\s*[-－— ]\s*(\d+)\s*$")


class ExcelFormatError(ValueError):
    """Excel 内容不符合约定格式。"""


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_yes(value) -> bool:
    return _text(value) in ("是", "Y", "y", "YES", "yes", "True", "true", "1")


def parse_config_sheet(ws) -> dict:
    """解析 config sheet，返回 {subjects, timelines, extras}。"""
    subjects: List[dict] = []
    row = 2
    while True:
        code = _text(ws.cell(row=row, column=1).value)
        name = _text(ws.cell(row=row, column=2).value)
        if not code and not name:
            break
        if row > 500:
            break
        try:
            code_int = int(float(code)) if code else row - 1
        except ValueError:
            raise ExcelFormatError("config!A%d 的科目代号必须是数字：%r" % (row, code))
        subjects.append({
            "code": code_int,
            "name": name,
            "short_name": _text(ws.cell(row=row, column=3).value),
            "outdoor": _is_yes(ws.cell(row=row, column=4).value),
        })
        row += 1

    if not subjects:
        raise ExcelFormatError("config sheet 未找到任何科目（B2 起填写科目列表）")

    # 时间线：E(5)..Z(26)，每条三列；行2=名称，行3=代号（可选），行4起=条目
    timelines: Dict[str, dict] = {}
    col = 5
    auto_index = 0
    while col + 2 <= 26 and auto_index < 7:
        name = _text(ws.cell(row=2, column=col).value)
        code = _text(ws.cell(row=3, column=col).value)
        entries = []
        r = 4
        while r <= 100:
            start = _text(ws.cell(row=r, column=col).value)
            duration = ws.cell(row=r, column=col + 1).value
            default_cls = ws.cell(row=r, column=col + 2).value
            if start == "" and duration is None and default_cls is None:
                break
            if start == "" or duration in (None, ""):
                raise ExcelFormatError(
                    "config!%s%d 时间线条目不完整（需要开始时间与持续分钟）"
                    % (ws.cell(row=r, column=col).coordinate, r)
                )
            entries.append({
                "start": start,
                "duration": int(float(duration)),
                "default_subject": _text(default_cls),
            })
            r += 1
        if not name and not code and not entries:
            break  # 空列：时间线区域结束
        auto_index += 1
        key = code or str(auto_index)
        if not name:
            name = "时间线" + key
        timelines[key] = {"code": key, "name": name, "entries": entries}
        col += 3

    # 杂项：AB(28) 名称，AC(29) 内容
    extras: Dict[str, str] = {}
    r = 2
    while r <= 500:
        key = _text(ws.cell(row=r, column=28).value)
        value = ws.cell(row=r, column=29).value
        if not key and value is None:
            break
        if key:
            extras[key] = "" if value is None else str(value)
        r += 1

    return {"subjects": subjects, "timelines": timelines, "extras": extras}


def parse_grade_sheet(ws) -> dict:
    """解析年级 sheet，返回 {classes: {class_name: [(day, period, subject_code)]}}。"""
    # 班级名：行 1，B1 起
    class_names: List[str] = []
    col = 2
    while col <= 200:
        name = _text(ws.cell(row=1, column=col).value)
        if not name:
            break
        class_names.append(name)
        col += 1
    if not class_names:
        raise ExcelFormatError("年级 sheet %r 的行 1 从 B1 起未找到班级名称" % ws.title)

    # 课位：A 列 "周几-第几节"
    classes = {name: [] for name in class_names}
    r = 2
    while r <= 1000:
        slot = _text(ws.cell(row=r, column=1).value)
        if not slot:
            break
        m = _SLOT_RE.match(slot)
        if not m:
            raise ExcelFormatError(
                "sheet %r 的 A%d 课位标记格式应为 \"周几-第几节\"，实际：%r"
                % (ws.title, r, slot)
            )
        day_name = m.group(1).replace("周天", "周日")
        day = DAY_ORDER[day_name]
        period = int(m.group(2))
        for i, cname in enumerate(class_names):
            raw = ws.cell(row=r, column=2 + i).value
            if raw is None or _text(raw) == "":
                continue
            try:
                code = int(float(_text(raw)))
            except ValueError:
                raise ExcelFormatError(
                    "sheet %r 的 %s%d 科目代号必须是数字，实际：%r"
                    % (ws.title, ws.cell(row=r, column=2 + i).coordinate, r, raw)
                )
            classes[cname].append({"day": day, "period": period, "subject_code": code})
        r += 1

    return {"classes": classes}


def parse_workbook(fileobj) -> dict:
    """解析整个工作簿。

    返回:
        {
          "subjects": [...], "timelines": {...}, "extras": {...},
          "grades": [{"name": sheet名, "classes": {班级名: [课位...]}}],
          "subject_codes": set(合法科目代号),
        }
    """
    wb = load_workbook(fileobj, read_only=False, data_only=True)
    config_name = None
    for title in wb.sheetnames:
        if title.strip().lower() == "config":
            config_name = title
            break
    if config_name is None:
        raise ExcelFormatError("缺少名为 config 的 sheet")

    cfg = parse_config_sheet(wb[config_name])
    subject_codes = {s["code"] for s in cfg["subjects"]}

    grades = []
    for title in wb.sheetnames:
        if title == config_name:
            continue
        result = parse_grade_sheet(wb[title])
        grades.append({"name": title.strip(), "classes": result["classes"]})

    if not grades:
        raise ExcelFormatError("工作簿中没有任何年级 sheet")

    # 校验科目代号
    for grade in grades:
        for cname, cells in grade["classes"].items():
            for cell in cells:
                if cell["subject_code"] not in subject_codes:
                    raise ExcelFormatError(
                        "年级 %r 班级 %r 引用了不存在的科目代号 %s"
                        % (grade["name"], cname, cell["subject_code"])
                    )

    cfg["grades"] = grades
    cfg["subject_codes"] = subject_codes
    return cfg
